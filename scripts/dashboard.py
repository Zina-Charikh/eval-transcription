"""
dashboard.py — SubQuality pipeline complet
Génère :
  - frames/whisper/  : frames annotées Whisper (mots erronés en rouge)
  - frames/cc/       : frames annotées CC YouTube (mots erronés en rouge)
  - dashboard.png
  - SubQuality.xlsx  (4 onglets dont CC YouTube)
  - wildmoka_clips.json

Usage : python3 dashboard.py
"""

import json, re, os
import numpy as np
import pandas as pd
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec
import matplotlib.patches as mpatches
from matplotlib.patches import FancyBboxPatch
from jiwer import wer, cer
try:
    from PIL import Image as PILImage, ImageDraw, ImageFont
    PIL_OK = True
except ImportError:
    PIL_OK = False
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

try:
    import cv2
    CV2_OK = True
except ImportError:
    CV2_OK = False
    print("⚠️  pip install opencv-python")

# ══════════════════════════════════════════════════════════════════
# CONFIG
# ══════════════════════════════════════════════════════════════════

TRANSCRIPTION  = "../data/transcription.json"
REFERENCE      = "../data/reference.txt"
SRT            = "../data/france24.fr.srt"
VIDEO          = "../france24.mp4"
EVAL_RESULTS   = "../outputs/eval_results.json"
OUT_PNG        = "../outputs/dashboard.png"
OUT_XLSX       = "../outputs/SubQuality.xlsx"
OUT_JSON       = "../outputs/wildmoka_clips.json"
OUT_W_FRAMES   = "../frames/whisper"
OUT_C_FRAMES   = "../frames/cc"

os.makedirs(OUT_W_FRAMES, exist_ok=True)
os.makedirs(OUT_C_FRAMES, exist_ok=True)

COLORS_Q = {
    "Parfait":    "#00C897",
    "Bon":        "#FFC107",
    "Acceptable": "#FF7043",
    "À corriger": "#E53935",
}
PAL = {
    "bg":     "#0D0F1A", "card":  "#161929",
    "accent": "#00C8FF", "red":   "#FF6B6B",
    "yellow": "#FFD166", "text":  "#E8EAF6",
    "grid":   "#1E2340", "muted": "#8892B0",
}

# ══════════════════════════════════════════════════════════════════
# 1. CHARGEMENT
# ══════════════════════════════════════════════════════════════════
print("📂 Chargement des fichiers...")

with open(TRANSCRIPTION, encoding="utf-8") as f:
    whisper_data = json.load(f)

whisper_segs = whisper_data["segments"]

with open(REFERENCE, encoding="utf-8") as f:
    ref_lines = [l.strip() for l in f if l.strip()]

# Charger eval_results
if os.path.exists(EVAL_RESULTS):

    with open(EVAL_RESULTS, encoding="utf-8") as f:
        ev = json.load(f)

    # Nouveau format JSON
    wh_wer_g = ev["global"]["whisper"]["wer_norm"]
    wh_cer_g = ev["global"]["whisper"]["cer"]

    cc_wer_g = ev["global"]["cc_youtube"]["wer_norm"]
    cc_cer_g = ev["global"]["cc_youtube"]["cer"]

    print("   ✅ eval_results.json chargé")

else:
    print("   ⚠️ Lance d'abord evaluate.py")

    cc_wer_g = 0.0
    cc_cer_g = 0.0
    wh_wer_g = 0.0
    wh_cer_g = 0.0


n = min(len(whisper_segs), len(ref_lines))

whisper_segs = whisper_segs[:n]
ref_lines    = ref_lines[:n]
print(f"   ✅ {n} segments alignés")

# ══════════════════════════════════════════════════════════════════
# 2. ALIGNEMENT CC YOUTUBE PAR SEGMENT
# ══════════════════════════════════════════════════════════════════

# Charge les CC YouTube alignés depuis eval_results.json (calculés par evaluate.py)
if os.path.exists(EVAL_RESULTS) and 'ev' in dir():
    cc_segments = [r.get("CC YouTube", "") for r in ev["segments"][:n]]
    print(f"   ✅ CC YouTube chargés ({sum(1 for c in cc_segments if c)} non vides)")
else:
    cc_segments = [""] * n
    print("   ⚠️  CC segments vides — lance evaluate.py d'abord")

# ══════════════════════════════════════════════════════════════════
# 3. MÉTRIQUES PAR SEGMENT
# ══════════════════════════════════════════════════════════════════

print("\n🔬 Calcul WER/CER par segment...")

def fmt(s): return f"{int(s)//60:02d}:{int(s)%60:02d}"

def quality_label(w):
    if w == 0:    return "Parfait"
    elif w < 10:  return "Bon"
    elif w < 25:  return "Acceptable"
    else:         return "À corriger"

def explain(ref, stt, w_val, c_val):
    reasons = []
    r2 = re.sub(r'[^\w\s]', '', ref.lower())
    h2 = re.sub(r'[^\w\s]', '', stt.lower())
    w2 = round(wer(r2, h2) * 100, 1)
    if w_val > 0 and w2 == 0:
        return "Seule la ponctuation diffère — sens identique"
    if w_val > 0 and w2 < w_val * 0.4:
        reasons.append("Principalement ponctuation/majuscules")
    noms = ["Tabarot","Flixbus","Ormesson","Météo-France",
            "Seine-et-Marne","A63","Île-de-France"]
    errs = [nm for nm in noms
            if nm.lower() in ref.lower()
            and nm.lower().split()[0] not in stt.lower()]
    if errs:
        reasons.append(f"Nom propre : {', '.join(errs[:2])}")
    diff = len(stt.split()) - len(ref.split())
    if diff > 1:    reasons.append(f"{diff} mots ajoutés")
    elif diff < -1: reasons.append(f"{abs(diff)} mots manquants")
    if w_val >= 40: reasons.append("Phrase fortement altérée")
    elif w_val >= 25: reasons.append("Plusieurs mots mal reconnus")
    if not reasons:
        if w_val == 0:    return "Transcription parfaite"
        elif w_val < 10:  return "Légère variation, sens préservé"
        else:             return "Écart modéré"
    note = (f" | CER {c_val}% bas car erreurs partielles"
            if w_val > 15 and c_val < w_val * 0.4 else "")
    return " + ".join(reasons) + note

records = []
for i, (seg, ref, cc) in enumerate(zip(whisper_segs, ref_lines, cc_segments)):
    stt    = seg["text"].strip()
    w_val  = round(wer(ref, stt) * 100, 1)
    c_val  = round(cer(ref, stt) * 100, 1)
    wc_val = round(wer(ref, cc)  * 100, 1) if cc else 0.0
    cc_val = round(cer(ref, cc)  * 100, 1) if cc else 0.0
    label  = quality_label(w_val)
    lbl_cc = quality_label(wc_val)

    records.append({
        "ID":            i + 1,
        "Début":         fmt(seg["start"]),
        "Fin":           fmt(seg["end"]),
        "start_s":       seg["start"],
        "end_s":         seg["end"],
        "Référence":     ref,
        "STT Whisper":   stt,
        "WER Whisper":   w_val,
        "CER Whisper":   c_val,
        "Qualité Whisper": label,
        "CC YouTube":    cc,
        "WER CC":        wc_val,
        "CER CC":        cc_val,
        "Qualité CC":    lbl_cc,
        "Explication":   explain(ref, stt, w_val, c_val),
        "Wildmoka":      ("🔴 À corriger" if w_val >= 25
                          else ("⚠️ Vérifier" if w_val > 0 else "✅ OK")),
    })

df = pd.DataFrame(records)
gw  = round(df["WER Whisper"].mean(), 1)
gc  = round(df["CER Whisper"].mean(), 1)
nbp = (df["WER Whisper"] == 0).sum()
nbc = (df["WER Whisper"] >= 25).sum()
print(f"   WER Whisper : {gw}%  |  WER CC YouTube : {cc_wer_g}%")
print(f"   Clips Wildmoka : {nbc}/{n}")

# ══════════════════════════════════════════════════════════════════
# 4. FRAMES VIDÉO — mots erronés en rouge, REF en bas
# ══════════════════════════════════════════════════════════════════

print("\n🎥 Extraction des frames vidéo...")

def mark_errors(ref, hyp):
    """
    Compare ref et hyp mot à mot.
    Retourne une liste de (mot, est_erroné) pour hyp.
    """
    ref_words = ref.lower().split()
    hyp_words = hyp.split()
    result = []
    ri = 0
    for hw in hyp_words:
        hw_clean = re.sub(r'[^\w]', '', hw.lower())
        if ri < len(ref_words) and hw_clean == re.sub(r'[^\w]', '', ref_words[ri]):
            result.append((hw, False))
            ri += 1
        else:
            result.append((hw, True))
    return result

def hex_to_rgb(h):
    h = h.lstrip("#")
    return tuple(int(h[i:i+2], 16) for i in (0, 2, 4))

def draw_subtitle_band(img_pil, draw, words_errors, y_top, label,
                        label_color_hex, bg_color_hex, font_sm, font_md,
                        img_w, img_h, line_h=28):
    """
    Dessine un bandeau de sous-titre PIL avec mots erronés en rouge.
    Mesure la largeur réelle de chaque mot via font.getbbox().
    """
    if not words_errors:
        return

    # Découpage en lignes (largeur max = 85% de l'image)
    MAX_W = int(img_w * 0.82)
    LABEL_W = int(img_w * 0.13)
    lines, cur_line, cur_w = [], [], 0
    for word, is_err in words_errors:
        bb = font_md.getbbox(word + " ")
        ww = bb[2] - bb[0]
        if cur_w + ww > MAX_W and cur_line:
            lines.append(cur_line)
            cur_line, cur_w = [(word, is_err, ww)], ww
        else:
            cur_line.append((word, is_err, ww))
            cur_w += ww
    if cur_line:
        lines.append(cur_line)
    lines = lines[:2]

    band_h = len(lines) * line_h + 14
    bg_rgb = hex_to_rgb(bg_color_hex)
    # Bandeau semi-transparent
    overlay = PILImage.new("RGBA", (img_w, band_h), (*bg_rgb, 210))
    img_pil.paste(PILImage.new("RGB", (img_w, band_h), bg_rgb),
                  (0, y_top), overlay)

    # Label source
    lr, lg, lb = hex_to_rgb(label_color_hex)
    draw.text((8, y_top + 5), label, font=font_sm, fill=(lr, lg, lb))

    # Mots
    for li, line_words in enumerate(lines):
        x = LABEL_W
        y = y_top + 5 + li * line_h
        for word, is_err, ww in line_words:
            color = (255, 70, 70) if is_err else (255, 255, 255)
            draw.text((x, y), word + " ", font=font_md, fill=color)
            x += ww
            if x > img_w - 10:
                break

def load_fonts(base_size):
    paths = [
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        "/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf",
        "/usr/share/fonts/truetype/freefont/FreeSans.ttf",
    ]
    bold_paths = [
        "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
        "/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf",
        "/usr/share/fonts/truetype/freefont/FreeSansBold.ttf",
    ]
    font_md = font_sm = font_bd = None
    for p, pb in zip(paths, bold_paths):
        try:
            font_md = ImageFont.truetype(p,  base_size)
            font_sm = ImageFont.truetype(p,  max(11, base_size - 2))
            font_bd = ImageFont.truetype(pb, base_size + 2)
            break
        except Exception:
            continue
    if font_md is None:
        font_md = font_sm = font_bd = ImageFont.load_default()
    return font_md, font_sm, font_bd

def generate_frame(frame_rgb, seg, ref, stt, source_label,
                   source_color, out_path):
    """
    Frame annotée claire :
      - Bandeau pleine largeur en haut  : badge source + WER
      - Bandeau pleine largeur (milieu) : STT avec mots erronés en rouge
      - Bandeau pleine largeur en bas   : REF en vert
    Design aéré, texte lisible, pas de chevauchement.
    """
    if not stt or not stt.strip():
        stt = "—"

    wer_val = round(wer(ref, stt) * 100, 1)
    col_hex = COLORS_Q[quality_label(wer_val)]
    col_rgb = hex_to_rgb(col_hex)

    img_h, img_w = frame_rgb.shape[:2]
    # Taille de police proportionnelle à la largeur image
    base_size = max(15, img_w // 26)
    font_md, font_sm, font_bd = load_fonts(base_size)

    img_pil = PILImage.fromarray(frame_rgb).convert("RGB")
    draw    = ImageDraw.Draw(img_pil)

    BAND_PAD   = 10          # padding intérieur vertical
    SIDE_PAD   = 14          # padding gauche
    LINE_H     = base_size + 8
    ALPHA_DARK = 200          # opacité fond sombre

    def draw_band(y_top, bg_rgb, lines_data, height=None):
        """
        Dessine un bandeau fond semi-transparent + lignes de texte.
        lines_data : liste de listes de (texte, couleur_rgb)
        """
        n_lines = len(lines_data)
        band_h  = height or (n_lines * LINE_H + 2 * BAND_PAD)
        # Fond semi-transparent via blend
        overlay = PILImage.new("RGB", (img_w, band_h), bg_rgb)
        img_pil.paste(overlay, (0, y_top))
        # Trait de séparation fin
        draw.line([(0, y_top), (img_w, y_top)], fill=(255,255,255,80), width=1)
        d = ImageDraw.Draw(img_pil)
        y = y_top + BAND_PAD
        for line in lines_data:
            x = SIDE_PAD
            for text, color in line:
                d.text((x, y), text, font=font_md, fill=color)
                bb = font_md.getbbox(text + " ")
                x += bb[2] - bb[0]
            y += LINE_H
        return band_h

    def wrap_words(words_colors, max_w):
        """Découpe (mot, couleur) en lignes ne dépassant pas max_w pixels."""
        lines, cur, cur_w = [], [], 0
        for word, color in words_colors:
            bb = font_md.getbbox(word + " ")
            ww = bb[2] - bb[0]
            if cur_w + ww > max_w and cur:
                lines.append(cur)
                cur, cur_w = [(word, color)], ww
            else:
                cur.append((word, color))
                cur_w += ww
        if cur:
            lines.append(cur)
        return lines[:3]  # max 3 lignes

    MAX_TXT_W = img_w - 2 * SIDE_PAD

    # ── BANDEAU 1 : Badge source + WER (fond couleur qualité) ────────
    badge_h = base_size + 2 * BAND_PAD
    overlay = PILImage.new("RGB", (img_w, badge_h), col_rgb)
    img_pil.paste(overlay, (0, 0))
    d = ImageDraw.Draw(img_pil)
    badge_txt = f"{source_label}  ·  Seg.{seg['ID']:02d}  ·  {seg['Début']}  ·  WER {wer_val}%"
    d.text((SIDE_PAD, BAND_PAD), badge_txt, font=font_bd, fill=(255, 255, 255))

    # ── BANDEAU 2 : STT avec erreurs (fond sombre, sous le badge) ────
    stt_errors  = mark_errors(ref, stt)
    words_color = [(w, (255,80,80) if err else (230,230,230))
                   for w, err in stt_errors]
    stt_lines   = wrap_words(words_color, MAX_TXT_W)
    stt_band_h  = len(stt_lines) * LINE_H + 2 * BAND_PAD
    y_stt = badge_h + 2
    overlay2 = PILImage.new("RGB", (img_w, stt_band_h), (20, 20, 40))
    img_pil.paste(overlay2, (0, y_stt))
    d = ImageDraw.Draw(img_pil)
    # Petit label source en haut à gauche du bandeau
    d.text((SIDE_PAD, y_stt + 3), source_label + " :", font=font_sm,
           fill=(160, 200, 255))
    y_txt = y_stt + BAND_PAD
    label_w = font_sm.getbbox(source_label + " :  ")[2]
    for li, line in enumerate(stt_lines):
        x = SIDE_PAD + (label_w if li == 0 else 0)
        for word, color in line:
            d.text((x, y_txt), word, font=font_md, fill=color)
            x += font_md.getbbox(word + " ")[2]
        y_txt += LINE_H

    # ── BANDEAU 3 : REF en bas (fond vert foncé) ─────────────────────
    ref_words   = [(w, (200, 255, 200)) for w in ref.split()]
    ref_lines_w = wrap_words(ref_words, MAX_TXT_W)
    ref_band_h  = len(ref_lines_w) * LINE_H + 2 * BAND_PAD
    y_ref = img_h - ref_band_h
    overlay3 = PILImage.new("RGB", (img_w, ref_band_h), (0, 50, 20))
    img_pil.paste(overlay3, (0, y_ref))
    d = ImageDraw.Draw(img_pil)
    d.text((SIDE_PAD, y_ref + 3), "REF :", font=font_sm, fill=(100, 220, 140))
    ref_label_w = font_sm.getbbox("REF :  ")[2]
    y_txt = y_ref + BAND_PAD
    for li, line in enumerate(ref_lines_w):
        x = SIDE_PAD + (ref_label_w if li == 0 else 0)
        for word, color in line:
            d.text((x, y_txt), word, font=font_md, fill=color)
            x += font_md.getbbox(word + " ")[2]
        y_txt += LINE_H

    img_pil.save(out_path)

frame_paths_w = []
frame_paths_c = []

if CV2_OK and os.path.exists(VIDEO):
    cap = cv2.VideoCapture(VIDEO)
    fps = cap.get(cv2.CAP_PROP_FPS) or 25
    print(f"   Vidéo ouverte — FPS : {fps}")

    # Frames Whisper — tous les segments WER >= 25% (sans limite)
    bad_w = df[df["WER Whisper"] >= 25]
    for _, seg in bad_w.iterrows():
        mid_s = (seg["start_s"] + seg["end_s"]) / 2
        cap.set(cv2.CAP_PROP_POS_FRAMES, int(mid_s * fps))
        ret, frame = cap.read()
        if not ret: continue
        frame_rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        out_path = os.path.join(OUT_W_FRAMES,
                                f"whisper_seg{seg['ID']:02d}.png")
        generate_frame(frame_rgb, seg, seg["Référence"],
                       seg["STT Whisper"], "Whisper", "#00C8FF", out_path)
        frame_paths_w.append(out_path)
        print(f"   ✅ Whisper frame Seg.{seg['ID']:02d} @ {mid_s:.1f}s")

    # Frames CC YouTube (segments WER CC >= 10%)
    bad_c = df[df["WER CC"] >= 10].head(4)
    if len(bad_c) == 0:
        # CC très bon, on prend quand même les pires
        bad_c = df.nlargest(4, "WER CC")
    for _, seg in bad_c.iterrows():
        mid_s = (seg["start_s"] + seg["end_s"]) / 2
        cap.set(cv2.CAP_PROP_POS_FRAMES, int(mid_s * fps))
        ret, frame = cap.read()
        if not ret: continue
        frame_rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        out_path = os.path.join(OUT_C_FRAMES,
                                f"cc_seg{seg['ID']:02d}.png")
        generate_frame(frame_rgb, seg, seg["Référence"],
                       seg["CC YouTube"], "CC YouTube", "#00C897", out_path)
        frame_paths_c.append(out_path)
        print(f"   ✅ CC YouTube frame Seg.{seg['ID']:02d} @ {mid_s:.1f}s")

    cap.release()
elif not os.path.exists(VIDEO):
    print(f"   ⚠️  Vidéo introuvable : {VIDEO}")

# ══════════════════════════════════════════════════════════════════
# 5. EXPORT JSON WILDMOKA
# ══════════════════════════════════════════════════════════════════

print("\n🎬 Export JSON Wildmoka...")

wildmoka = {
    "source":          "France 24 — JT Épisode Neige",
    "pipeline":        "SubQuality v1.0",
    "model_stt":       "openai/whisper-base",
    "global_wer":      gw, "global_cer": gc,
    "cc_youtube_wer":  cc_wer_g,
    "total_segments":  n,
    "clips_to_review": int(nbc),
    "clips":           [],
}
for _, r in df[df["WER Whisper"] >= 25].iterrows():
    wildmoka["clips"].append({
        "clip_id":       f"F24_SEG_{r['ID']:02d}",
        "timestamp_in":  r["Début"],
        "timestamp_out": r["Fin"],
        "start_s":       r["start_s"],
        "end_s":         r["end_s"],
        "wer_score":     r["WER Whisper"],
        "cer_score":     r["CER Whisper"],
        "quality":       r["Qualité Whisper"],
        "priority":      "HIGH" if r["WER Whisper"] >= 40 else "MEDIUM",
        "action":        "REVIEW_REQUIRED",
        "wildmoka_tag":  "subtitle_quality_alert",
        "reference":     r["Référence"],
        "stt_output":    r["STT Whisper"],
        "explication":   r["Explication"],
    })

with open(OUT_JSON, "w", encoding="utf-8") as f:
    json.dump(wildmoka, f, ensure_ascii=False, indent=2)
print(f"   ✅ {nbc} clips → wildmoka_clips.json")

# ══════════════════════════════════════════════════════════════════
# 6. DASHBOARD PNG
# ══════════════════════════════════════════════════════════════════

print("\n📊 Génération dashboard PNG...")

plt.rcParams.update({
    "figure.facecolor": PAL["bg"],   "axes.facecolor":  PAL["card"],
    "axes.edgecolor":   PAL["grid"], "axes.labelcolor": PAL["text"],
    "xtick.color":      PAL["text"], "ytick.color":     PAL["text"],
    "text.color":       PAL["text"], "grid.color":      PAL["grid"],
    "font.family":      "DejaVu Sans", "font.size": 9,
})

fig = plt.figure(figsize=(22, 16), facecolor=PAL["bg"])
fig.suptitle(
    f"SubQuality — France 24 × Whisper × CC YouTube × Wildmoka\n"
    f"WER Whisper : {gw}%  |  WER CC YouTube : {cc_wer_g}%  |  "
    f"{nbc}/{n} clips à corriger",
    fontsize=15, fontweight="bold", color=PAL["accent"], y=0.98
)

gs = gridspec.GridSpec(3, 4, figure=fig, hspace=0.48, wspace=0.38)

# ── WER Whisper par segment ──────────────────────────────────────
ax1 = fig.add_subplot(gs[0, :3])
bars = ax1.bar(df["ID"], df["WER Whisper"],
               color=[COLORS_Q[q] for q in df["Qualité Whisper"]],
               width=0.7, zorder=3)
ax1.axhline(gw,       color=PAL["accent"], linestyle="--", lw=1.5,
            label=f"WER moyen Whisper : {gw}%", zorder=4)
ax1.axhline(cc_wer_g, color="#00C897",    linestyle="-.", lw=1.5,
            label=f"WER CC YouTube : {cc_wer_g}%", zorder=4)
ax1.axhline(25,       color=PAL["yellow"], linestyle=":", lw=1.2,
            label="Seuil Wildmoka : 25%", zorder=4)
ax1.set_xlabel("Segment n°", fontsize=9)
ax1.set_ylabel("WER (%)", fontsize=9)
ax1.set_title("WER par segment — Whisper base — France 24",
              fontsize=11, fontweight="bold", pad=8)
ax1.legend(loc="upper right", fontsize=8, framealpha=0.3)
ax1.grid(axis="y", alpha=0.3, zorder=0)
ax1.set_xticks(df["ID"])
for bar, row in zip(bars, df.itertuples()):
    if row._8 > 0:
        ax1.text(bar.get_x() + bar.get_width()/2,
                 bar.get_height() + 0.4,
                 f"{row._8}%", ha="center", va="bottom",
                 fontsize=7, color="white")

# ── KPIs ─────────────────────────────────────────────────────────
ax_k = fig.add_subplot(gs[0, 3])
ax_k.set_xlim(0, 1); ax_k.set_ylim(0, 1); ax_k.axis("off")
ax_k.set_title("Indicateurs clés", fontsize=10, fontweight="bold", pad=8)
for i, (lbl, val, col) in enumerate([
    ("WER Whisper",    f"{gw}%",       PAL["accent"]),
    ("CER Whisper",    f"{gc}%",       "#A78BFA"),
    ("WER CC YouTube", f"{cc_wer_g}%", "#00C897"),
    ("CER CC YouTube", f"{cc_cer_g}%", "#34D399"),
    ("Clips Wildmoka", str(nbc),       PAL["red"]),
]):
    y = 0.90 - i * 0.19
    ax_k.add_patch(plt.Rectangle((0.03, y-0.07), 0.94, 0.14,
                                  facecolor=PAL["bg"], edgecolor=col, lw=1.5))
    ax_k.text(0.10, y, lbl, va="center", fontsize=8, color=PAL["text"])
    ax_k.text(0.93, y, val, va="center", ha="right",
              fontsize=11, fontweight="bold", color=col)

# ── Distribution qualité ─────────────────────────────────────────
ax2 = fig.add_subplot(gs[1, 0])
q_order  = ["Parfait", "Bon", "Acceptable", "À corriger"]
q_counts = [df["Qualité Whisper"].value_counts().get(q, 0) for q in q_order]
wedges, texts, autotexts = ax2.pie(
    q_counts, labels=q_order,
    colors=[COLORS_Q[q] for q in q_order],
    autopct=lambda p: f"{p:.0f}%" if p > 0 else "",
    startangle=90, pctdistance=0.72,
    wedgeprops=dict(width=0.55, edgecolor=PAL["bg"], linewidth=2)
)
for t in texts:      t.set_fontsize(8); t.set_color(PAL["text"])
for at in autotexts: at.set_fontsize(8); at.set_color("white")
ax2.set_title("Distribution qualité\n(Whisper)", fontsize=10,
              fontweight="bold", pad=6)

# ── Whisper vs CC YouTube ────────────────────────────────────────
ax3 = fig.add_subplot(gs[1, 1:3])
cats   = ["WER global", "CER global"]
vals_w = [gw, gc]
vals_c = [cc_wer_g, cc_cer_g]
x = np.arange(len(cats)); bw = 0.35
ax3.bar(x - bw/2, vals_w, bw, color=PAL["accent"],
        label="Whisper base", alpha=0.85, zorder=3)
ax3.bar(x + bw/2, vals_c, bw, color="#00C897",
        label="CC YouTube",   alpha=0.85, zorder=3)
for i, (vw, vc) in enumerate(zip(vals_w, vals_c)):
    ax3.text(i - bw/2, vw + 0.3, f"{vw}%",
             ha="center", fontsize=9, fontweight="bold", color=PAL["accent"])
    ax3.text(i + bw/2, vc + 0.3, f"{vc}%",
             ha="center", fontsize=9, fontweight="bold", color="#00C897")
ax3.set_xticks(x); ax3.set_xticklabels(cats)
ax3.set_ylabel("Taux d'erreur (%)", fontsize=9)
ax3.set_title("Comparaison globale Whisper vs CC YouTube",
              fontsize=10, fontweight="bold", pad=6)
ax3.legend(fontsize=9, framealpha=0.3)
ax3.grid(axis="y", alpha=0.3)

# ── WER vs CER scatter ───────────────────────────────────────────
ax4 = fig.add_subplot(gs[1, 3])
ax4.scatter(df["WER Whisper"], df["CER Whisper"],
            c=[COLORS_Q[q] for q in df["Qualité Whisper"]],
            s=70, alpha=0.9, zorder=3, edgecolors="white", linewidths=0.5)
for _, r in df.iterrows():
    ax4.annotate(f"S{r['ID']}", (r["WER Whisper"], r["CER Whisper"]),
                 textcoords="offset points", xytext=(3, 2),
                 fontsize=6.5, color=PAL["text"])
ax4.set_xlabel("WER (%)", fontsize=9)
ax4.set_ylabel("CER (%)", fontsize=9)
ax4.set_title("Corrélation\nWER / CER", fontsize=10, fontweight="bold", pad=6)
ax4.grid(alpha=0.3)
ax4.legend(handles=[mpatches.Patch(color=COLORS_Q[l], label=l)
                    for l in q_order
                    if df["Qualité Whisper"].value_counts().get(l, 0) > 0],
           fontsize=6.5, framealpha=0.3)

# ── Timeline WER ─────────────────────────────────────────────────
ax5 = fig.add_subplot(gs[2, :2])
ax5.plot(df["ID"], df["WER Whisper"], color=PAL["accent"], lw=1.8, zorder=3)
ax5.fill_between(df["ID"], df["WER Whisper"],
                 alpha=0.12, color=PAL["accent"])
for _, r in df.iterrows():
    ax5.plot(r["ID"], r["WER Whisper"], "o", markersize=6,
             color=COLORS_Q[r["Qualité Whisper"]], zorder=4)
ax5.axhline(25, color=PAL["yellow"], linestyle=":", lw=1.2,
            label="Seuil 25%", alpha=0.7)
ax5.set_xlabel("Segment n°", fontsize=9)
ax5.set_ylabel("WER (%)", fontsize=9)
ax5.set_title("Évolution du WER — chronologie vidéo",
              fontsize=10, fontweight="bold", pad=6)
ax5.legend(fontsize=8, framealpha=0.3)
ax5.grid(alpha=0.3); ax5.set_xticks(df["ID"])

# ── Clips Wildmoka ───────────────────────────────────────────────
ax6 = fig.add_subplot(gs[2, 2:])
ax6.set_xlim(0, 1); ax6.set_ylim(0, 1); ax6.axis("off")
ax6.set_title("Wildmoka — Clips marqués", fontsize=10,
              fontweight="bold", pad=6)
for i, (_, r) in enumerate(df[df["WER Whisper"] >= 25].head(7).iterrows()):
    y = 0.92 - i * 0.135
    col = COLORS_Q[r["Qualité Whisper"]]
    prio = "HAUTE" if r["WER Whisper"] >= 40 else "MOYENNE"
    ax6.add_patch(plt.Rectangle((0.01, y-0.06), 0.98, 0.12,
                                 facecolor=PAL["bg"], edgecolor=col, lw=1.5))
    ax6.text(0.04, y+0.01,
             f"Seg.{r['ID']:02d} · {r['Début']}→{r['Fin']}",
             fontsize=7.5, va="center", color=PAL["text"], fontweight="bold")
    ax6.text(0.04, y-0.03, r["Explication"][:55],
             fontsize=6.5, va="center", color=PAL["muted"])
    ax6.text(0.97, y, f"WER {r['WER Whisper']}% | {prio}",
             ha="right", va="center", fontsize=8,
             fontweight="bold", color=col)

plt.savefig(OUT_PNG, dpi=150, bbox_inches="tight", facecolor=PAL["bg"])
plt.close()
print("   ✅ dashboard.png généré")

# ══════════════════════════════════════════════════════════════════
# 7. EXCEL (4 onglets avec CC YouTube)
# ══════════════════════════════════════════════════════════════════

print("\n📊 Génération Excel...")

def fill(h):   return PatternFill("solid", fgColor=h.lstrip("#"))
def fnt(bold=False, size=10, color="E8EAF6"):
    return Font(name="Arial", bold=bold, size=size, color=color)
def ctr(): return Alignment(horizontal="center", vertical="center",
                             wrap_text=True)
def brd():
    s = Side(style="thin", color="2E3560")
    return Border(left=s, right=s, top=s, bottom=s)

C_D = "#0D0F1A"; C_N = "#161929"; C_H = "#1E2340"
wb = Workbook(); wb.remove(wb.active)

def make_header_row(ws, row_n, headers, col_colors=None):
    ws.row_dimensions[row_n].height = 22
    for i, h in enumerate(headers, 1):
        c = get_column_letter(i)
        ws[f"{c}{row_n}"] = h
        ws[f"{c}{row_n}"].font = fnt(bold=True, size=9)
        ws[f"{c}{row_n}"].alignment = ctr()
        ws[f"{c}{row_n}"].fill = fill(C_H)
        ws[f"{c}{row_n}"].border = brd()

def make_title(ws, cell, text, color="00C8FF"):
    ws[cell] = text
    ws[cell].font = Font(name="Arial", bold=True, size=13, color=color)
    ws[cell].alignment = ctr()
    ws[cell].fill = fill(C_N)

# ── Onglet 1 : Vue d'ensemble ────────────────────────────────────
ws0 = wb.create_sheet("Vue d'ensemble")
for r in ws0.iter_rows(1, 30, 1, 6):
    for c in r: c.fill = fill(C_D)
ws0.column_dimensions["A"].width = 2
for i, w in enumerate([28, 16, 16, 14, 36], 2):
    ws0.column_dimensions[get_column_letter(i)].width = w

ws0.merge_cells("B1:F1"); ws0.row_dimensions[1].height = 34
make_title(ws0, "B1", "SubQuality — Vue d'ensemble comparative")

overview = [
    ("Métrique",        "Whisper base",       "CC YouTube",      "Meilleur",  "Interprétation"),
    ("WER global (%)",  f"{gw}%",             f"{cc_wer_g}%",   "CC YouTube","CC YouTube plus précis — Google Speech entraîné sur YouTube"),
    ("CER global (%)",  f"{gc}%",             f"{cc_cer_g}%",   "CC YouTube","Même tendance au niveau des caractères"),
    ("Granularité",     "Par segment+timestamp","Texte global",  "Whisper",   "Whisper localise l'erreur à la seconde — indispensable pour Wildmoka"),
    ("Wildmoka JSON",   "Oui — timestamps",   "Non",            "Whisper",   "Export JSON avec timestamp_in/out pour chaque clip à corriger"),
    ("Amélioration",    "large-v3 → ~14% WER","Non modifiable", "Whisper",   "Changer le modèle STT sans changer le pipeline"),
]

make_header_row(ws0, 4, overview[0])
for ri, row in enumerate(overview[1:], 5):
    ws0.row_dimensions[ri].height = 42
    bg = C_N if ri % 2 == 0 else "#12152A"
    for ci, val in enumerate(row, 2):
        cell = ws0[f"{get_column_letter(ci)}{ri}"]
        cell.value = val; cell.fill = fill(bg); cell.border = brd()
        cell.font = Font(name="Arial", size=9,
                         color="00C8FF" if ci == 2 else
                         ("00C897" if ci == 5 and "CC" in str(val)
                          else ("00C8FF" if ci == 5 and "Whisper" in str(val)
                                else "E8EAF6")),
                         bold=(ci == 2))
        cell.alignment = Alignment(
            horizontal="center" if ci < 6 else "left",
            vertical="center", wrap_text=True)

# ── Onglet 2 : Analyse Whisper ───────────────────────────────────
ws1 = wb.create_sheet("Analyse Whisper")
for r in ws1.iter_rows(1, 80, 1, 11):
    for c in r: c.fill = fill(C_D)
for i, w in enumerate([2, 5, 8, 8, 40, 40, 9, 9, 16, 35, 14], 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

ws1.merge_cells("A1:K1"); ws1.row_dimensions[1].height = 30
make_title(ws1, "A1", "Analyse Whisper — segment par segment vs Référence")

hdrs_w = ["ID", "Début", "Fin", "Référence", "STT Whisper",
          "WER %", "CER %", "Qualité", "Explication", "Wildmoka"]
make_header_row(ws1, 3, hdrs_w)

for ri, (_, row) in enumerate(df.iterrows(), 4):
    ws1.row_dimensions[ri].height = 45
    bg = C_N if ri % 2 == 0 else "#12152A"
    qc = COLORS_Q.get(row["Qualité Whisper"], "#888888").lstrip("#")
    vals = [row["ID"], row["Début"], row["Fin"], row["Référence"],
            row["STT Whisper"], row["WER Whisper"], row["CER Whisper"],
            row["Qualité Whisper"], row["Explication"], row["Wildmoka"]]
    for ci, val in enumerate(vals, 1):
        cell = ws1[f"{get_column_letter(ci)}{ri}"]
        cell.value = val; cell.fill = fill(bg); cell.border = brd()
        cell.font = Font(name="Arial", size=8.5,
                         color=qc if ci in [8, 9] else "E8EAF6",
                         bold=(ci == 8))
        cell.alignment = Alignment(
            horizontal="center" if ci < 5 else "left",
            vertical="center", wrap_text=True)

# Frames Whisper dans l'Excel
row_frames = 4 + len(df) + 2
ws1[f"A{row_frames}"] = "Frames vidéo Whisper — mots erronés en rouge (OpenCV)"
ws1[f"A{row_frames}"].font = Font(name="Arial", bold=True, size=11, color="00C8FF")
ws1.row_dimensions[row_frames].height = 18
col_off = 1
for fp in frame_paths_w[:3]:
    if os.path.exists(fp):
        img = XLImage(fp); img.width, img.height = 380, 280
        ws1.add_image(img, f"{get_column_letter(col_off)}{row_frames+1}")
        col_off += 5

if os.path.exists(OUT_PNG):
    img = XLImage(OUT_PNG); img.width, img.height = 980, 700
    ws1.add_image(img, f"A{row_frames + 22}")

# ── Onglet 3 : Analyse CC YouTube ───────────────────────────────
ws2 = wb.create_sheet("Analyse CC YouTube")
for r in ws2.iter_rows(1, 80, 1, 10):
    for c in r: c.fill = fill(C_D)
for i, w in enumerate([2, 5, 8, 8, 40, 40, 9, 9, 16, 14], 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

ws2.merge_cells("A1:J1"); ws2.row_dimensions[1].height = 30
make_title(ws2, "A1",
           f"Analyse CC YouTube — WER global {cc_wer_g}% vs Référence",
           color="00C897")

ws2.merge_cells("A2:J2"); ws2.row_dimensions[2].height = 35
ws2["A2"] = (
    f"CC YouTube = sous-titres automatiques générés par Google Speech via YouTube. "
    f"WER global {cc_wer_g}% vs {gw}% pour Whisper. "
    f"Plus précis globalement mais sans timestamps par segment → impossible d'intégrer à Wildmoka."
)
ws2["A2"].font = Font(name="Arial", size=9, italic=True, color="8892B0")
ws2["A2"].alignment = Alignment(horizontal="left", vertical="center",
                                wrap_text=True)
ws2["A2"].fill = fill("#12152A")

hdrs_c = ["ID", "Début", "Fin", "Référence", "CC YouTube",
          "WER %", "CER %", "Qualité CC"]
make_header_row(ws2, 4, hdrs_c)

for ri, (_, row) in enumerate(df.iterrows(), 5):
    ws2.row_dimensions[ri].height = 45
    bg = C_N if ri % 2 == 0 else "#12152A"
    qc = COLORS_Q.get(row["Qualité CC"], "#888888").lstrip("#")
    vals = [row["ID"], row["Début"], row["Fin"], row["Référence"],
            row["CC YouTube"], row["WER CC"], row["CER CC"],
            row["Qualité CC"]]
    for ci, val in enumerate(vals, 1):
        cell = ws2[f"{get_column_letter(ci)}{ri}"]
        cell.value = val; cell.fill = fill(bg); cell.border = brd()
        cell.font = Font(name="Arial", size=8.5,
                         color=qc if ci == 8 else "E8EAF6",
                         bold=(ci == 8))
        cell.alignment = Alignment(
            horizontal="center" if ci < 5 else "left",
            vertical="center", wrap_text=True)

# Frames CC dans l'Excel
row_fc = 5 + len(df) + 2
ws2[f"A{row_fc}"] = "Frames vidéo CC YouTube — mots erronés en rouge"
ws2[f"A{row_fc}"].font = Font(name="Arial", bold=True, size=11, color="00C897")
ws2.row_dimensions[row_fc].height = 18
col_off2 = 1
for fp in frame_paths_c[:3]:
    if os.path.exists(fp):
        img = XLImage(fp); img.width, img.height = 380, 280
        ws2.add_image(img, f"{get_column_letter(col_off2)}{row_fc+1}")
        col_off2 += 5

# ── Onglet 4 : Wildmoka ─────────────────────────────────────────
ws3 = wb.create_sheet("Wildmoka — Clips")
for r in ws3.iter_rows(1, 50, 1, 9):
    for c in r: c.fill = fill(C_D)
for i, w in enumerate([2, 18, 10, 10, 9, 9, 10, 12, 45], 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

ws3.merge_cells("A1:I1"); ws3.row_dimensions[1].height = 28
make_title(ws3, "A1",
           "Export Wildmoka — Clips subtitle_quality_alert (WER ≥ 25%)")

ws3.merge_cells("A2:I2"); ws3.row_dimensions[2].height = 55
ws3["A2"] = (
    "WILDMOKA : plateforme de clipping média (France 24, Arte, BFM TV). "
    "Ce tableau simule l'export vers l'API Wildmoka : chaque clip contient "
    "le timestamp exact du problème dans la vidéo, permettant à l'éditeur "
    "de retrouver le moment, corriger le sous-titre et valider avant diffusion."
)
ws3["A2"].font = Font(name="Arial", size=9, italic=True, color="8892B0")
ws3["A2"].alignment = Alignment(horizontal="left", vertical="center",
                                wrap_text=True)
ws3["A2"].fill = fill("#12152A")

wm_hdrs = ["clip_id", "timestamp_in", "timestamp_out",
           "WER %", "CER %", "Priorité", "Action", "Explication"]
make_header_row(ws3, 4, wm_hdrs)

for ri, clip in enumerate(wildmoka["clips"], 5):
    ws3.row_dimensions[ri].height = 40
    bg = C_N if ri % 2 == 0 else "#12152A"
    col = "E53935" if clip["priority"] == "HIGH" else "FF7043"
    vals = [clip["clip_id"], clip["timestamp_in"], clip["timestamp_out"],
            clip["wer_score"], clip["cer_score"],
            "HAUTE" if clip["priority"] == "HIGH" else "MOYENNE",
            "RÉVISION REQUISE", clip["explication"]]
    for ci, val in enumerate(vals, 1):
        cell = ws3[f"{get_column_letter(ci)}{ri}"]
        cell.value = val; cell.fill = fill(bg); cell.border = brd()
        cell.font = Font(name="Arial", size=9,
                         color=col if ci == 6 else "E8EAF6",
                         bold=(ci == 6))
        cell.alignment = Alignment(
            horizontal="center" if ci < 8 else "left",
            vertical="center", wrap_text=True)

wb.save(OUT_XLSX)
print("   ✅ SubQuality.xlsx généré (4 onglets)")

# ══════════════════════════════════════════════════════════════════
# RÉSUMÉ
# ══════════════════════════════════════════════════════════════════

print(f"""
{'='*55}
  PIPELINE TERMINÉ !
{'='*55}
  dashboard.png
  SubQuality.xlsx         (4 onglets)
  wildmoka_clips.json     ({nbc} clips)
  frames/whisper/         ({len(frame_paths_w)} frames Whisper)
  frames/cc/              ({len(frame_paths_c)} frames CC YouTube)
{'='*55}
  WER Whisper    : {gw}%
  WER CC YouTube : {cc_wer_g}%
  Clips Wildmoka : {nbc}/{n}
{'='*55}
""")
